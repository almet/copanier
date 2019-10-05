from dataclasses import fields as get_fields

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from .models import Product, Producer


def summary_for_products(wb, title, delivery, total=None, products=None):
    if products == None:
        products = delivery.products
    if total == None:
        total = delivery.total

    ws = wb.create_sheet(title)
    ws.append(
        ["ref", "produit", "prix unitaire", "quantité commandée", "unité", "total"]
    )
    for product in products:
        wanted = delivery.product_wanted(product)
        if not wanted:
            continue
        ws.append(
            [
                product.ref,
                str(product),
                product.price,
                wanted,
                product.unit,
                round(product.price * wanted, 2),
            ]
        )
    ws.append(["", "", "", "", "Total", total])


def summary(delivery, producers=None):
    wb = Workbook()
    wb.remove(wb.active)
    if not producers:
        producers = delivery.producers
    for producer in producers:
        summary_for_products(
            wb,
            producer,
            delivery,
            total=delivery.total_for_producer(producer),
            products=delivery.get_products_by(producer),
        )

    return save_virtual_workbook(wb)


def full(delivery):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{delivery.name} {delivery.from_date.date()}"
    headers = ["ref", "produit", "prix"] + [e for e in delivery.orders] + ["total"]
    headers.insert(1, "producer")
    ws.append(headers)
    for product in delivery.products:
        row = [product.ref, str(product), product.price]
        row.insert(1, product.producer)
        for order in delivery.orders.values():
            wanted = order.products.get(product.ref)
            row.append(wanted.quantity if wanted else 0)
        row.append(delivery.product_wanted(product))
        ws.append(row)
    footer = (
        ["Total", "", ""]
        + [round(o.total(delivery.products), 2) for o in delivery.orders.values()]
        + [round(delivery.total, 2)]
    )
    footer.insert(1, "")

    ws.append(footer)
    return save_virtual_workbook(wb)


def products(delivery):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{delivery.name} produits"
    product_fields = [f.name for f in get_fields(Product)]
    ws.append(product_fields)
    for product in delivery.products:
        ws.append([getattr(product, field) for field in product_fields])

    producer_sheet = wb.create_sheet(f"producteur⋅ice⋅s et référent⋅e⋅s")
    producer_fields = [f.name for f in get_fields(Producer)]
    producer_sheet.append(producer_fields)
    for producer in delivery.producers.values():
        producer_sheet.append([getattr(producer, field) for field in producer_fields])

    return save_virtual_workbook(wb)
